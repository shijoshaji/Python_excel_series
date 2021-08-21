# Working with Sheets

from openpyxl import load_workbook

# NOTE: loading the excel we need
wb = load_workbook(filename='sampleData.xlsx')

# NOTE: get the sheetnames from the excel we read
print(wb.sheetnames)  # OUTPUT: ['Sheet1', 'Sheet2']

# NOTE: shows which sheet is currently active
print(wb.active)  # OUTPUT: <Worksheet "Sheet1">

# NOTE: we can assign which sheet to be activated, it starts from left to right from the index 0,1 so on...
wb.active = 0
print(wb.active)  # OUTPUT: <Worksheet "Sheet1">
wb.active = 1
print(wb.active)  # OUTPUT: <Worksheet "Sheet2">

# NOTE: above we saw we can access via index location, now lets access the sheet with the sheet name
sheet = wb['Sheet2']
print(sheet)  # OUTPUT: <Worksheet "Sheet2">
print(sheet.title)  # OUTPUT: Sheet2
