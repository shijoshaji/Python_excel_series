# Retrieving Cell Values

from openpyxl import load_workbook

# NOTE: loading the excel we need
wb = load_workbook(filename='sampleData.xlsx')

sheet = wb['Sheet1']
# NOTE: from active sheet we are trying to fetch the value from cell B3
cell_coordinates = sheet['B3']

# NOTE: fetch value, row & column for the cell coordinates
print(cell_coordinates.value)  # OUTPUT: Jojo
print(cell_coordinates.row)  # OUTPUT: 3
print(cell_coordinates.column)  # OUTPUT: 2
print(cell_coordinates.coordinate)  # OUTPUT: B3

# NOTE: What if we fetch the empty cell value?
print(sheet['B9'].value)  # OUTPUT: None

# NOTE: Return Value using cell
print(sheet.cell(row=2, column=2).value)  # OUTPUT: Shijo
