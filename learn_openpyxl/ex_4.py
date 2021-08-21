# Retrieving Multiple Values

from openpyxl import load_workbook

# NOTE: loading the excel we need
wb = load_workbook(filename='sampleData.xlsx')

sheet = wb['Sheet2']

# NOTE: fetches all the 'A' colum that has data
# OUTPUT: (<Cell 'Sheet2'.A1>, <Cell 'Sheet2'.A2>, <Cell 'Sheet2'.A3>, <Cell 'Sheet2'.A4>, <Cell 'Sheet2'.A5>, <Cell 'Sheet2'.A6>)
print(sheet['A'])

# NOTE: fetches range of columns without index
print(sheet['A:C'])
'''
OUTOUT:
((<Cell 'Sheet2'.A1>, <Cell 'Sheet2'.A2>, <Cell 'Sheet2'.A3>, <Cell 'Sheet2'.A4>, <Cell 'Sheet2'.A5>, <Cell 'Sheet2'.A6>), 
(<Cell 'Sheet2'.B1>, <Cell 'Sheet2'.B2>, <Cell 'Sheet2'.B3>, <Cell 'Sheet2'.B4>, <Cell 'Sheet2'.B5>, <Cell 'Sheet2'.B6>), 
(<Cell 'Sheet2'.C1>, <Cell 'Sheet2'.C2>, <Cell 'Sheet2'.C3>, <Cell 'Sheet2'.C4>, <Cell 'Sheet2'.C5>, <Cell 'Sheet2'.C6>))
'''

# NOTE: fetches range of columns with index
print(sheet['1:3'])
'''
OUTPUT:
((<Cell 'Sheet2'.A1>, <Cell 'Sheet2'.B1>, <Cell 'Sheet2'.C1>), 
(<Cell 'Sheet2'.A2>, <Cell 'Sheet2'.B2>, <Cell 'Sheet2'.C2>), 
(<Cell 'Sheet2'.A3>, <Cell 'Sheet2'.B3>, <Cell 'Sheet2'.C3>))
'''

# fetch row & column objects
for row in sheet.rows:
    print(row)

'''
OUTPUT:
(<Cell 'Sheet2'.A1>, <Cell 'Sheet2'.B1>, <Cell 'Sheet2'.C1>)
(<Cell 'Sheet2'.A2>, <Cell 'Sheet2'.B2>, <Cell 'Sheet2'.C2>)
(<Cell 'Sheet2'.A3>, <Cell 'Sheet2'.B3>, <Cell 'Sheet2'.C3>)
(<Cell 'Sheet2'.A4>, <Cell 'Sheet2'.B4>, <Cell 'Sheet2'.C4>)
(<Cell 'Sheet2'.A5>, <Cell 'Sheet2'.B5>, <Cell 'Sheet2'.C5>)
(<Cell 'Sheet2'.A6>, <Cell 'Sheet2'.B6>, <Cell 'Sheet2'.C6>)
'''

for col in sheet.columns:
    print(col)

'''
OUTPUT:
(<Cell 'Sheet2'.A1>, <Cell 'Sheet2'.A2>, <Cell 'Sheet2'.A3>, <Cell 'Sheet2'.A4>, <Cell 'Sheet2'.A5>, <Cell 'Sheet2'.A6>)
(<Cell 'Sheet2'.B1>, <Cell 'Sheet2'.B2>, <Cell 'Sheet2'.B3>, <Cell 'Sheet2'.B4>, <Cell 'Sheet2'.B5>, <Cell 'Sheet2'.B6>)
(<Cell 'Sheet2'.C1>, <Cell 'Sheet2'.C2>, <Cell 'Sheet2'.C3>, <Cell 'Sheet2'.C4>, <Cell 'Sheet2'.C5>, <Cell 'Sheet2'.C6>)
'''

# Show values only
for row in sheet.iter_rows(values_only=True):
    print(row)

'''
OUTPUT:
('id', 'fruit', 'quantity')
(1, 'Apple', 2)
(2, 'Orange', 5)
(3, 'Pineapple', 3)
(4, 'Banana', 6)
(5, 'Mango', 1)
'''

for col in sheet.iter_cols(values_only=True):
    print(col)

'''
OUTPUT:
('id', 1, 2, 3, 4, 5)
('fruit', 'Apple', 'Orange', 'Pineapple', 'Banana', 'Mango')
('quantity', 2, 5, 3, 6, 1)
'''
