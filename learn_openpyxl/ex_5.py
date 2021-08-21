# Converting data into Python structures
import json
from openpyxl import load_workbook

# NOTE: loading the excel we need
wb = load_workbook(filename='sampleData.xlsx')

sheet = wb['Sheet2']
#  empty dictionary to keep values from excel
books = {}

for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
    book_id = row[0]
    book = {
        'Fruit': row[1],
        'Qty': row[2]
    }
    books[book_id] = book

print(json.dumps(books, indent=3))

'''
OUTPUT:
{
   "1": {
      "Fruit": "Apple",
      "Qty": 2
   },
   "2": {
      "Fruit": "Orange",
      "Qty": 5
   },
   "3": {
      "Fruit": "Pineapple",
      "Qty": 3
   },
   "4": {
      "Fruit": "Banana",
      "Qty": 6
   },
   "5": {
      "Fruit": "Mango",
      "Qty": 1
   }
}
'''
